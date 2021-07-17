#!/usr/bin/env python3

from io import StringIO
import csv
import re
import sys
import argparse
from contextlib import ExitStack
from typing import Iterable, List, Tuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class DuplicateDataError(Exception):
    def __init__(self, old, new, message):
        self.old = old
        self.new = new
        self.message = message

        super().__init__(f'{message} old: {old} -> new: {new}')


def is_blank_row(row: List[str]) -> bool:
    return not row or all(cell in ('', 'NR') for cell in row[1:])


def merge_tables(str_tables: Iterable[str]) -> str:
    data = dict()
    tables = [list(csv.reader(table.splitlines())) for table in str_tables]

    for row in tables[0]:
        if row:
            data[row[0]] = row

    for table in tables:
        for row in table:
            if not is_blank_row(row):
                if row[0] in data:
                    if not is_blank_row(data[row[0]]) and data[row[0]] != row:
                        raise DuplicateDataError(data[row[0]], row, f'Duplicate data for {row[0]}.')
                data[row[0]] = row

    out = StringIO()
    writer = csv.writer(out)
    for row in tables[0]:
        if not row:
            continue
        best_row = data[row[0]]
        writer.writerow(best_row)

    return out.getvalue()


_RE_FOO_RESULTS_TABLE = re.compile(r'"(?P<tbl_name>\S+ Results) Table"')


def extract_tables(contents: str) -> Iterable[Tuple[str, str]]:
    for m in _RE_FOO_RESULTS_TABLE.finditer(contents):
        tbl_start = contents.find('\n\n', m.end()) + 1
        tbl_end = contents.find('\n\n', tbl_start)
        yield (m['tbl_name'], contents[tbl_start:tbl_end])


def main(files, out: str):
    wb = Workbook()
    files = [f.read() for f in files]
    xy = list(extract_tables(files[0]))
    tbls = map(extract_tables, files)
    for tbl_group in zip(*tbls):
        assert len(set(name for name, _ in tbl_group)) == 1
        ws = wb.create_sheet(tbl_group[0][0])

        str_tables = (tbl for _, tbl in tbl_group)
        merged = merge_tables(str_tables)
        for row in csv.reader(merged.splitlines()):
            ws.append(row)
            for i, _ in enumerate(row):
                ws.column_dimensions[get_column_letter(i + 1)].bestFit = True

    wb.remove(wb.active)
    wb.save(out)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Merges multiple CPU2017 exec time csv results together')
    parser.add_argument('-o', '--output', required=True, help='Where to write the output file')
    parser.add_argument('csvs', nargs='+', help='The files to merge')

    args = parser.parse_args()

    with ExitStack() as stack:
        files = [stack.enter_context(open(f, 'r')) for f in args.csvs]

        main(files, args.output)
