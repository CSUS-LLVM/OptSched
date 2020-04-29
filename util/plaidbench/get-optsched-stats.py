#!/usr/bin/python3
'''
**********************************************************************************
Description:    This script is meant to be used with the OptSched scheduler and
                the run-plaidbench.sh script. This script will extract stats
                about how our OptSched scheduler is doing from the log files
                generated from the run-plaidbench.sh script.
Author:         Vang Thao
Last Update:    December 30, 2019
**********************************************************************************

OUTPUT:
    This script takes in data from plaidbench runs and output a single spreadsheet.
        Spreadsheet 1: optsched-stats.xlsx

Requirements:
    - python3
    - pip3
    - openpyxl (sreadsheet module, installed using pip3)

HOW TO USE:
    1.) Run a plaidbench benchmarks with run-plaidbench.sh to generate a
        directory containing the results for the run.
    2.) Pass the path to the directory as an input to this script with
        the -i option.

Example:
    ./get-optsched-stats.py -i /home/tom/plaidbench-optsched-01/

    where plaidbench-optsched-01/ contains
        densenet121
        densenet169
        ...
        ...
'''

import os       # Used for scanning directories, getting paths, and checking files.
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import argparse

REGEX_DAG_INFO = re.compile(r'Processing DAG (.*) with (\d+) insts and max latency (\d+)')
REGEX_LIST_OPTIMAL = re.compile(r'list schedule (.?)* is optimal')
REGEX_COST_IMPROV = re.compile(r'cost imp=(\d+).')
REGEX_OPTIMAL = re.compile(r'The schedule is optimal')
REGEX_PASS_NUM = re.compile(r'End of (.*) pass through')

# Contains all of the stats
benchStats = {}
passStats = {}
passes = ['first', 'second', 'third']

# List of benchmark names
benchmarks = [
    'densenet121',
    'densenet169',
    'densenet201',
    'inception_resnet_v2',
    'inception_v3',
    'mobilenet',
    'nasnet_large',
    'nasnet_mobile',
    'resnet50',
    'vgg16',
    'vgg19',
    'xception',
    'imdb_lstm',
]

# List of stats that can be initialized to 0
statsProcessed = [
    'TotalProcessed',
    'EnumCnt',
    'OptImpr',
    'OptNotImpr',
    'TimeoutImpr',
    'TimeoutNotImpr',
    'TimeoutCnt',
    'TotalInstr'
]

def initializePassStats(dictToInitialize):
    for stat in statsProcessed:
        dictToInitialize[stat] = 0
    dictToInitialize['AverageSizeToEnum'] = -1.0
    dictToInitialize['LargestOptimalRegion'] = -1
    dictToInitialize['LargestImprovedRegion'] = -1

def parseStats(inputFolder):
    # Initialize pass stats collection variables
    for x in passes:
        passStats[x] = {}
        initializePassStats(passStats[x])

    # Begin stats collection for this run
    for bench in benchmarks:
        # Get the path to the log file
        currentPath = os.path.join(inputFolder, bench)
        currentLogFile = os.path.join(currentPath, bench + '.log')

        # First check if log file exists.
        if os.path.exists(currentLogFile):
            benchStats[bench] = {}
            # Open log file if it exists.
            with open(currentLogFile) as file:
                # Contain the stats for this benchmark
                stats = {}
                for x in passes:
                    stats[x] = {}
                    initializePassStats(stats[x])

                log = file.read()
                blocks = log.split('********** Opt Scheduling **********')[1:]
                for block in blocks:
                    # Get pass num, if none is found then
                    # use third as default.
                    getPass = REGEX_PASS_NUM.search(block)
                    if getPass:
                        passNum = getPass.group(1)
                    else:
                        passNum = "third"

                    stats[passNum]['TotalProcessed'] += 1

                    # If our enumerator was called then
                    # record stats for it.
                    if 'Enumerating' in block:
                        stats[passNum]['EnumCnt'] += 1
                        # Get cost
                        searchCost = REGEX_COST_IMPROV.search(block)
                        cost = int(searchCost.group(1))

                        # Get DAG stats
                        dagInfo = REGEX_DAG_INFO.search(block)
                        numOfInstr = int(dagInfo.group(2))
                        stats[passNum]['TotalInstr'] += numOfInstr

                        if REGEX_OPTIMAL.search(block):
                            # Optimal and improved
                            if cost > 0:
                                stats[passNum]['OptImpr'] += 1
                                if (numOfInstr > stats[passNum]['LargestImprovedRegion']):
                                    stats[passNum]['LargestImprovedRegion'] = numOfInstr
                            # Optimal but not improved
                            elif cost == 0:
                                stats[passNum]['OptNotImpr'] += 1
                            if (numOfInstr > stats[passNum]['LargestOptimalRegion']):
                                stats[passNum]['LargestOptimalRegion'] = numOfInstr
                        elif 'timedout' in block:
                            # Timeout and improved
                            if cost > 0:
                                stats[passNum]['TimeoutImpr'] += 1
                                if numOfInstr > stats[passNum]['LargestImprovedRegion']:
                                    stats[passNum]['LargestImprovedRegion'] = numOfInstr
                            # Timeout but not improved
                            elif cost == 0:
                                stats[passNum]['TimeoutNotImpr'] += 1
                            stats[passNum]['TimeoutCnt'] += 1


        # If the file doesn't exist, output error log.
        else:
            print('Cannot find log file for benchmark {}.'.format(bench))

        for passNum in passes:
            for stat in statsProcessed:
                passStats[passNum][stat] += stats[passNum][stat]
            if stats[passNum]['EnumCnt'] != 0:
                stats[passNum]['AverageSizeToEnum'] = float(stats[passNum]['TotalInstr'])/stats[passNum]['EnumCnt']
            if passStats[passNum]['LargestOptimalRegion'] < stats[passNum]['LargestOptimalRegion']:
                passStats[passNum]['LargestOptimalRegion'] = stats[passNum]['LargestOptimalRegion']
            if passStats[passNum]['LargestImprovedRegion'] < stats[passNum]['LargestImprovedRegion']:
                passStats[passNum]['LargestImprovedRegion'] = stats[passNum]['LargestImprovedRegion']

        benchStats[bench] = stats

    for passNum in passes:
        if passStats[passNum]['EnumCnt'] != 0:
            passStats[passNum]['AverageSizeToEnum'] = float(passStats[passNum]['TotalInstr'])/passStats[passNum]['EnumCnt']

def printStats():
    for passNum in passes:
        if passStats[passNum]['TotalProcessed'] == 0:
            continue

        # Third pass is just a default if the two-pass
        # implementation wasn't used and not an actual
        # 3 pass implementation.
        if passNum == 'third':
            print('first')
        else:
            print(passNum)

        for stat in passStats[passNum]:
            print('    {} : {}'.format(stat, passStats[passNum][stat]))

def writeBenchmarkNames(ws, row):
    for bench in benchmarks:
        ws['A' + str(row)] = bench
        row += 1
    ws['A' + str(row)] = 'Overall'
    ws['A' + str(row)].font = Font(bold=True)

def createSpreadsheets(output):
    if 'xls' not in output[-4:]:
        output += '.xlsx'

    # Create new excel worksheet
    wb = Workbook()

    # Grab the active worksheet
    ws = wb.active

    # Insert column titles
    ws['A1'] = 'Benchmarks'
    ws['A1'].font = Font(bold=True)
    col = 'B'
    row = 1
    ws[col + str(row)] = 'Benchmark Stats'
    row = 2
    ws[col+str(row)] = 'Regions processed'
    ws[chr(ord(col)+1)+str(row)] = 'Passed to B&B'
    ws[chr(ord(col)+2)+str(row)] = 'Optimal and improved'
    ws[chr(ord(col)+3)+str(row)] = 'Optimal and not improved'
    ws[chr(ord(col)+4)+str(row)] = 'Timed out and improved'
    ws[chr(ord(col)+5)+str(row)] = 'Timed out and not improved'
    ws[chr(ord(col)+6)+str(row)] = 'Avg. Region size passed to B&B'
    ws[chr(ord(col)+7)+str(row)] = 'Largest optimal region'
    ws[chr(ord(col)+8)+str(row)] = 'Largest improved region'

    # Stats entry
    row = 3
    for passNum in passes:
        # Skip pass if there is no data.
        if passStats[passNum]['TotalProcessed'] == 0:
            continue

        # Identify each pass if data set is from 2-pass
        # scheduler.
        if not passNum == 'third':
            ws['A'+str(row-1)] = passNum.capitalize() + ' Pass'
            ws['A'+str(row-1)].font = Font(bold=True)

        writeBenchmarkNames(ws, row)

        # Write individual benchmark stats
        for bench in benchmarks:
            ws[col+str(row)] = benchStats[bench][passNum]['TotalProcessed']
            ws[col+str(row)].alignment = Alignment(horizontal='right')
            if benchStats[bench][passNum]['EnumCnt'] != 0:
                enumCntPcnt = float(benchStats[bench][passNum]['EnumCnt']) / benchStats[bench][passNum]['TotalProcessed'] * 100.0
                ws[chr(ord(col)+1)+str(row)] = str(benchStats[bench][passNum]['EnumCnt']) + ' ({:.2f}%)'.format(enumCntPcnt)
                ws[chr(ord(col)+1)+str(row)].alignment = Alignment(horizontal='right')

                enumCntPcnt = float(benchStats[bench][passNum]['OptImpr']) / benchStats[bench][passNum]['EnumCnt'] * 100.0
                ws[chr(ord(col)+2)+str(row)] = str(benchStats[bench][passNum]['OptImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
                ws[chr(ord(col)+2)+str(row)].alignment = Alignment(horizontal='right')

                enumCntPcnt = float(benchStats[bench][passNum]['OptNotImpr']) / benchStats[bench][passNum]['EnumCnt'] * 100.0
                ws[chr(ord(col)+3)+str(row)] = str(benchStats[bench][passNum]['OptNotImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
                ws[chr(ord(col)+3)+str(row)].alignment = Alignment(horizontal='right')

                enumCntPcnt = float(benchStats[bench][passNum]['TimeoutImpr']) / benchStats[bench][passNum]['EnumCnt'] * 100.0
                ws[chr(ord(col)+4)+str(row)] = str(benchStats[bench][passNum]['TimeoutImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
                ws[chr(ord(col)+4)+str(row)].alignment = Alignment(horizontal='right')

                enumCntPcnt = float(benchStats[bench][passNum]['TimeoutNotImpr']) / benchStats[bench][passNum]['EnumCnt'] * 100.0
                ws[chr(ord(col)+5)+str(row)] = str(benchStats[bench][passNum]['TimeoutNotImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
                ws[chr(ord(col)+5)+str(row)].alignment = Alignment(horizontal='right')

                ws[chr(ord(col)+6)+str(row)] = benchStats[bench][passNum]['AverageSizeToEnum']
                ws[chr(ord(col)+6)+str(row)].alignment = Alignment(horizontal='right')

                ws[chr(ord(col)+7)+str(row)] = benchStats[bench][passNum]['LargestOptimalRegion']
                ws[chr(ord(col)+7)+str(row)].alignment = Alignment(horizontal='right')

                ws[chr(ord(col)+8)+str(row)] = benchStats[bench][passNum]['LargestImprovedRegion']
                ws[chr(ord(col)+8)+str(row)].alignment = Alignment(horizontal='right')

            row += 1

        # Write overall stats
        ws[col+str(row)] = passStats[passNum]['TotalProcessed']
        enumCntPcnt = float(passStats[passNum]['EnumCnt']) / passStats[passNum]['TotalProcessed'] * 100.0
        ws[chr(ord(col)+1)+str(row)] = str(passStats[passNum]['EnumCnt']) + ' ({:.2f}%)'.format(enumCntPcnt)
        ws[chr(ord(col)+1)+str(row)].alignment = Alignment(horizontal='right')

        if passStats[passNum]['EnumCnt'] != 0:
            enumCntPcnt = float(passStats[passNum]['OptImpr']) / passStats[passNum]['EnumCnt'] * 100.0
            ws[chr(ord(col)+2)+str(row)] = str(passStats[passNum]['OptImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
            ws[chr(ord(col)+2)+str(row)].alignment = Alignment(horizontal='right')

            enumCntPcnt = float(passStats[passNum]['OptNotImpr']) / passStats[passNum]['EnumCnt'] * 100.0
            ws[chr(ord(col)+3)+str(row)] = str(passStats[passNum]['OptNotImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
            ws[chr(ord(col)+3)+str(row)].alignment = Alignment(horizontal='right')

            enumCntPcnt = float(passStats[passNum]['TimeoutImpr']) / passStats[passNum]['EnumCnt'] * 100.0
            ws[chr(ord(col)+4)+str(row)] = str(passStats[passNum]['TimeoutImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
            ws[chr(ord(col)+4)+str(row)].alignment = Alignment(horizontal='right')

            enumCntPcnt = float(passStats[passNum]['TimeoutNotImpr']) / passStats[passNum]['EnumCnt'] * 100.0
            ws[chr(ord(col)+5)+str(row)] = str(passStats[passNum]['TimeoutNotImpr']) + ' ({:.2f}%)'.format(enumCntPcnt)
            ws[chr(ord(col)+5)+str(row)].alignment = Alignment(horizontal='right')

            ws[chr(ord(col)+6)+str(row)] = passStats[passNum]['AverageSizeToEnum']
            ws[chr(ord(col)+6)+str(row)].alignment = Alignment(horizontal='right')

            ws[chr(ord(col)+7)+str(row)] = passStats[passNum]['LargestOptimalRegion']
            ws[chr(ord(col)+7)+str(row)].alignment = Alignment(horizontal='right')

            ws[chr(ord(col)+8)+str(row)] = passStats[passNum]['LargestImprovedRegion']
            ws[chr(ord(col)+8)+str(row)].alignment = Alignment(horizontal='right')

        # Prepare to write for next pass if there is any.
        row += 3

    wb.save(output)

def main(args):
    # Start stats collection
    parseStats(args.inputFolder)

    if args.verbose:
        printStats()

    if not args.disable:
        createSpreadsheets(args.output)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Script to extract OptSched stats', \
                                      formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('--verbose', '-v',
                        action='store_true', default=False,
                        dest='verbose',
                        help='Print the stats to terminal')

    parser.add_argument('--output', '-o',
                        default='optsched-stats',
                        dest='output',
                        help='Output spreadsheet filepath')

    parser.add_argument('--disable', '-d',
                        action='store_true', default=False,
                        dest='disable',
                        help='Disable spreadsheet output.')

    parser.add_argument('--input', '-i',
                        default='.',
                        dest='inputFolder',
                        help='The path to scan for benchmark directories')

    args = parser.parse_args()

    main(args)
