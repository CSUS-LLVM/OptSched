#!/usr/bin/python3
'''
Extract execution-time scores from SHOC runs. Will look for the results of each
individual benchmark in the "Logs" directory.
Ex: optsched-run-01/Logs/dev0_FFT.log

Since SHOC doesn't allow you to select only a subset of benchmarks to run, you
can use this command to run the level1 benchmarks we are interested in:
cd $SHOC/build
mkdir optsched-01 && cd optsched-01
mkdir shoc-fft && cd $_ && ../../bin/shocdriver -opencl -benchmark FFT -s 4 && cd .. && \
mkdir shoc-gemm && cd $_ && ../../bin/shocdriver -opencl -benchmark GEMM -s 4 && cd .. && \
mkdir shoc-md && cd $_ && ../../bin/shocdriver -opencl -benchmark MD -s 4 && cd .. && \
mkdir shoc-sort && cd $_ && ../../bin/shocdriver -opencl -benchmark Sort -s 4 && cd .. && \
mkdir shoc-spmv && cd $_ && ../../bin/shocdriver -opencl -benchmark Spmv -s 4 && cd .. && \
mkdir shoc-stencil2d && cd $_ && ../../bin/shocdriver -opencl -benchmark Stencil2D -s 4 && cd ..

Then you must copy the logs and err files over to the Logs directory.
mkdir Logs
cp shoc-*/Logs/*.err Logs
cp shoc-*/Logs/*.log Logs

Feed the result into this script
./extract-shoc-data.py optsched-run-01/
'''

import argparse
import os
import re
import statistics
from openpyxl import Workbook
from openpyxl.styles import Font

RE_BENCHMARK_NAME = re.compile('Running benchmark (.*)')
RE_BENCHMARK_RESULTS = re.compile('result for (.*):(\s)*(.*) (.*)')


# Name of the results.txt file
filenameShoc = 'shoc.log'

displayOrder = {
    'FFT': ['fft_sp', 'ifft_sp', 'fft_dp', 'ifft_dp'],
    'GEMM': ['sgemm_n', 'sgemm_t', 'dgemm_n', 'dgemm_t'],
    'MD': ['md_sp_flops', 'md_sp_bw', 'md_dp_flops', 'md_dp_bw'],
    'Sort': ['sort'],
    'Spmv': [
        'spmv_csr_scalar_sp',
        'spmv_csr_scalar_dp',
        'spmv_csr_scalar_pad_sp',
        'spmv_csr_scalar_pad_dp',
        'spmv_csr_vector_sp',
        'spmv_csr_vector_dp',
        'spmv_csr_vector_pad_sp',
        'spmv_csr_vector_pad_dp',
        'spmv_ellpackr_sp',
        'spmv_ellpackr_dp'
    ],
    'Stencil2D': ['stencil', 'stencil_dp'],
    # 'S3D': ['s3d', 's3d_dp'],
}

benchmarks = [
    'BusSpeedDownload',
    'BusSpeedReadback',
    'MaxFlops',
    'DeviceMemory',
    'KernelCompile',
    'QueueDelay',
    'BFS',
    'FFT',
    'GEMM',
    'MD',
    'MD5Hash',
    'Reduction',
    'Scan',
    'Sort',
    'Spmv',
    'Stencil2D',
    'Triad',
    'S3D'
]

dataFormat = [
    'test',
    'atts',
    'units',
    'median',
    'mean',
    'stddev',
    'min',
    'max',
    'trial0',
    'trial1',
    'trial2',
    'trial3',
    'trial4',
    'trial5',
    'trial6',
    'trial7',
    'trial8',
    'trial9',
]

queueDelayDataFormat = [
    'test',
    'atts',
    'units',
    'median',
    'mean',
    'stddev',
    'min',
    'max',
    'trial0',
    'trial1',
    'trial2',
]


def parseStats(inputFolder):
    # Get the path to the logs folder
    currentPath = os.path.join(inputFolder, 'Logs')
    stats = {
        # 'BusSpeedDownload'  : processBusSpeedDownload(currentPath),
        # 'BusSpeedReadback'  : processBusSpeedReadback(currentPath),
        # 'MaxFlops'          : processMaxFlops(currentPath),
        # 'DeviceMemory'      : processDeviceMemory(currentPath),
        # 'KernelCompile'     : processKernelCompile(currentPath),
        # 'QueueDelay'        : processQueueDelay(currentPath),
        # 'BFS'               : processBFS(currentPath),
        'FFT': processFTT(currentPath),
        'GEMM': processGEMM(currentPath),
        'MD': processMD(currentPath),
        # 'MD5Hash'           : processMD5Hash(currentPath),
        # 'Reduction'         : processReduction(currentPath),
        # 'Scan'              : processScan(currentPath),
        'Sort': processSort(currentPath),
        'Spmv': processSpmv(currentPath),
        'Stencil2D': processStencil2D(currentPath),
        # 'Triad'             : processTriad(currentPath),
        # 'S3D'               : processS3D(currentPath)
    }
    return stats


def printStats(stats):
    for benchmark in stats:
        print('{}'.format(benchmark))
        for test in stats[benchmark]:
            if '_pcie' in test:
                continue
            print('    {}'.format(test))
            print('      {}'.format(stats[benchmark][test]))


def createSpreadsheet(inputFolder, output, stats):
    if 'xls' not in output[-4:]:
        output += '.xlsx'

    # Create new excel worksheet
    wb = Workbook()

    # Grab the active worksheet
    ws = wb.active

    # Insert column titles
    ws['A1'] = 'Benchmarks'
    ws['A1'].font = Font(bold=True)
    ws['C1'] = inputFolder
    ws['C1'].font = Font(bold=True)
    ws['C2'] = 'Score'
    ws['D2'] = 'Units'
    ws['E2'] = 'RV'

    col = 'A'
    row = 3
    for benchmark in displayOrder:
        ws[col+str(row)] = benchmark

        for test in displayOrder[benchmark]:
            ws[chr(ord(col)+1)+str(row)] = test

            scores = [float(stats[benchmark][test]['trial9']),
                      float(stats[benchmark][test]['trial8']),
                      float(stats[benchmark][test]['trial7']),
                      float(stats[benchmark][test]['trial6']),
                      float(stats[benchmark][test]['trial5']),
                      ]

            median = statistics.median(scores)
            randomVar = (max(scores) - min(scores)) / min(scores)

            ws[chr(ord(col)+2)+str(row)] = median
            ws[chr(ord(col)+3)+str(row)] = stats[benchmark][test]['units']
            ws[chr(ord(col)+4)+str(row)] = randomVar
            row += 1

    wb.save(output)

# Input: Logs folder
# Output: One dictionarary


def processBusSpeedDownload(folder):
    stats = {}
    filename = 'dev0_BusSpeedDownload.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                # Process the line as a space delimited list
                curLine = line.split()
                if len(curLine) > 2:
                    # The line we want to check is usually the line with the
                    # biggest size.
                    if curLine[1] == '524288kB' and \
                            curLine[0] == 'DownloadSpeed':
                        # Insert the stats into a dictionary with its
                        # coressponding data format name as the key.
                        stats = {dataFormat[i]: curLine[i]
                                 for i in range(len(dataFormat))}
    else:
        print('Cannot find log file {}'.format(filename))

    return {'bspeed_download': stats}

# Input: Logs folder
# Output: One dictionarary, two entries:
# bspeed_download (type dict)
# bspeed_readback (type dict)


def processBusSpeedReadback(folder):
    stats = {}
    filename = 'dev0_BusSpeedReadback.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                # Process the line as a space delimited list
                curLine = line.split()
                if len(curLine) > 2:
                    # The line we want to check is usually the line with the
                    # biggest size.
                    if curLine[1] == '524288kB' and curLine[0] == 'ReadbackSpeed':
                        # Insert the stats into a dictionary with its coressponding
                        # data format name as the key.
                        stats = {dataFormat[i]: curLine[i]
                                 for i in range(len(dataFormat))}
    else:
        print('Cannot find log file {}'.format(filename))

    return {'bspeed_readback': stats}


def processMaxFlops(folder):
    bestSP = {}
    bestDP = {}
    filename = 'dev0_MaxFlops.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            # Track current state on whether to process data.
            dataStart = False
            maxFlopsSP = 0.0
            maxFlopsDP = 0.0
            curTest = ''

            for line in f:
                curLine = line.split()
                # There is a blank newline which indicates the end.
                # Check if we reached that newline.
                # Stop processing data if we reached end.
                if len(curLine) == 0:
                    dataStart = False

                # We are in the middle of the data block.
                # Process data.
                elif dataStart:
                    curStat = {dataFormat[i]: curLine[i]
                               for i in range(len(dataFormat))}
                    curTest = curStat['test'].split('-')[-1]

                    if curTest == 'SP':
                        if float(curStat['median']) > maxFlopsSP:
                            maxFlopsSP = float(curStat['max'])
                            bestSP = curStat
                    elif curTest == 'DP':
                        if float(curStat['median']) > maxFlopsDP:
                            maxFlopsDP = float(curStat['max'])
                            bestDP = curStat

                # Current line reached where the data is located
                # Start processing data
                elif curLine == dataFormat:
                    dataStart = True
    else:
        print('Cannot find log file {}'.format(filename))

    return {'maxspflops': bestSP, 'maxdpflops': bestDP}


def processDeviceMemory(folder):
    stats = {}
    filename = 'dev0_DeviceMemory.log'
    max_readGlobalMemoryCoalesced = -1.0
    max_readGlobalMemoryUnit = -1.0
    max_writeGlobalMemoryCoalesced = -1.0
    max_writeGlobalMemoryUnit = -1.0
    max_readLocalMemory = -1.0
    max_writeLocalMemory = -1.0
    max_TextureRepeatedRandomAccess = -1.0
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'readGlobalMemoryCoalesced':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_readGlobalMemoryCoalesced:
                            max_readGlobalMemoryCoalesced = float(
                                curStats['median'])
                            stats['gmem_readbw'] = curStats
                    elif curLine[0] == 'readGlobalMemoryUnit':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_readGlobalMemoryUnit:
                            max_readGlobalMemoryUnit = float(
                                curStats['median'])
                            stats['gmem_readbw_strided'] = curStats
                    elif curLine[0] == 'writeGlobalMemoryCoalesced':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_writeGlobalMemoryCoalesced:
                            max_writeGlobalMemoryCoalesced = float(
                                curStats['median'])
                            stats['gmem_writebw'] = curStats
                    elif curLine[0] == 'writeGlobalMemoryUnit':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_writeGlobalMemoryUnit:
                            max_writeGlobalMemoryUnit = float(
                                curStats['median'])
                            stats['gmem_writebw_strided'] = curStats
                    elif curLine[0] == 'readLocalMemory':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_readLocalMemory:
                            max_readLocalMemory = float(curStats['median'])
                            stats['lmem_readbw'] = curStats
                    elif curLine[0] == 'writeLocalMemory':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_writeLocalMemory:
                            max_writeLocalMemory = float(curStats['median'])
                            stats['lmem_writebw'] = curStats
                    elif curLine[0] == 'TextureRepeatedRandomAccess':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max_TextureRepeatedRandomAccess:
                            max_TextureRepeatedRandomAccess = float(
                                curStats['median'])
                            stats['tex_readbw'] = curStats
    return stats


def processKernelCompile(folder):
    stats = {}
    minCompileTime = 9999999.0
    filename = 'dev0_KernelCompile.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'BuildProgram':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) < minCompileTime:
                            minCompileTime = float(curStats['median'])
                            stats = curStats
    else:
        print('Cannot find log file {}'.format(filename))

    return {'ocl_kernel': stats}


def processQueueDelay(folder):
    stats = {}
    minDelay = 9999999.0
    filename = 'dev0_QueueDelay.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'SSDelay':
                        curStats = {queueDelayDataFormat[i]: curLine[i] for i in range(
                            len(queueDelayDataFormat))}
                        if float(curStats['median']) < minDelay:
                            minDelay = float(curStats['median'])
                            stats = curStats
    else:
        print('Cannot find log file {}'.format(filename))

    return {'ocl_queue': stats}


def processBFS(folder):
    stats = {}
    filename = 'dev0_BFS.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'BFS':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['bfs'] = curStats
                    elif curLine[0] == 'BFS_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['bfs_pcie'] = curStats
                    elif curLine[0] == 'BFS_teps':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['bfs_teps'] = curStats

    return stats


def processFTT(folder):
    stats = {}
    filename = 'dev0_FFT.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'SP-FFT':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['fft_sp'] = curStats
                    elif curLine[0] == 'SP-FFT_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['fft_sp_pcie'] = curStats
                    elif curLine[0] == 'SP-FFT-INV':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['ifft_sp'] = curStats
                    elif curLine[0] == 'SP-FFT-INV_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['ifft_sp_pcie'] = curStats
                    elif curLine[0] == 'DP-FFT':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['fft_dp'] = curStats
                    elif curLine[0] == 'DP-FFT_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['fft_dp_pcie'] = curStats
                    elif curLine[0] == 'DP-FFT-INV':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['ifft_dp'] = curStats
                    elif curLine[0] == 'DP-FFT-INV_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['ifft_dp_pcie'] = curStats

    return stats


def processGEMM(folder):
    stats = {}
    filename = 'dev0_GEMM.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'SGEMM-N':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['sgemm_n'] = curStats
                    elif curLine[0] == 'SGEMM-T':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['sgemm_t'] = curStats
                    elif curLine[0] == 'SGEMM-N_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['sgemm_n_pcie'] = curStats
                    elif curLine[0] == 'SGEMM-T_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['sgemm_t_pcie'] = curStats
                    elif curLine[0] == 'DGEMM-N':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['dgemm_n'] = curStats
                    elif curLine[0] == 'DGEMM-T':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['dgemm_t'] = curStats
                    elif curLine[0] == 'DGEMM-N_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['dgemm_n_pcie'] = curStats
                    elif curLine[0] == 'DGEMM-T_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['dgemm_t_pcie'] = curStats

    return stats


def processMD(folder):
    stats = {}
    filename = 'dev0_MD.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'MD-LJ':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_sp_flops'] = curStats
                    elif curLine[0] == 'MD-LJ-Bandwidth':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_sp_bw'] = curStats
                    elif curLine[0] == 'MD-LJ_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_sp_flops_pcie'] = curStats
                    elif curLine[0] == 'MD-LJ-Bandwidth_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_sp_bw_pcie'] = curStats
                    elif curLine[0] == 'MD-LJ-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_dp_flops'] = curStats
                    elif curLine[0] == 'MD-LJ-DP-Bandwidth':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_dp_bw'] = curStats
                    elif curLine[0] == 'MD-LJ-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_dp_flops_pcie'] = curStats
                    elif curLine[0] == 'MD-LJ-DP-Bandwidth_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md_dp_bw_pcie'] = curStats

    return stats


def processMD5Hash(folder):
    stats = {}
    filename = 'dev0_MD5Hash.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'MD5Hash':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['md5hash'] = curStats

    return stats


def processReduction(folder):
    stats = {}
    filename = 'dev0_Reduction.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'Reduction':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['reduction'] = curStats
                    elif curLine[0] == 'Reduction_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['reduction_pcie'] = curStats
                    elif curLine[0] == 'Reduction-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['reduction_dp'] = curStats
                    elif curLine[0] == 'Reduction-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['reduction_dp_pcie'] = curStats

    return stats


def processScan(folder):
    stats = {}
    filename = 'dev0_Scan.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'Scan':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['scan'] = curStats
                    elif curLine[0] == 'Scan_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['scan_pcie'] = curStats
                    elif curLine[0] == 'Scan-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['scan_dp'] = curStats
                    elif curLine[0] == 'Scan-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['scan_dp_pcie'] = curStats

    return stats


def processSort(folder):
    stats = {}
    filename = 'dev0_Sort.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'Sort-Rate':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['sort'] = curStats
                    elif curLine[0] == 'Sort-Rate_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['sort_pcie'] = curStats

    return stats


def processSpmv(folder):
    stats = {}
    filename = 'dev0_Spmv.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'CSR-Scalar-SP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_sp'] = curStats
                    elif curLine[0] == 'CSR-Scalar-SP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_sp_pcie'] = curStats
                    elif curLine[0] == 'CSR-Scalar-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_dp'] = curStats
                    elif curLine[0] == 'CSR-Scalar-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_dp_pcie'] = curStats
                    elif curLine[0] == 'Padded_CSR-Scalar-SP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_pad_sp'] = curStats
                    elif curLine[0] == 'Padded_CSR-Scalar-SP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_pad_sp_pcie'] = curStats
                    elif curLine[0] == 'Padded_CSR-Scalar-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_pad_dp'] = curStats
                    elif curLine[0] == 'Padded_CSR-Scalar-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_scalar_pad_dp_pcie'] = curStats
                    elif curLine[0] == 'CSR-Vector-SP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_sp'] = curStats
                    elif curLine[0] == 'CSR-Vector-SP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_sp_pcie'] = curStats
                    elif curLine[0] == 'CSR-Vector-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_dp'] = curStats
                    elif curLine[0] == 'CSR-Vector-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_dp_pcie'] = curStats
                    elif curLine[0] == 'Padded_CSR-Vector-SP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_pad_sp'] = curStats
                    elif curLine[0] == 'Padded_CSR-Vector-SP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_pad_sp_pcie'] = curStats
                    elif curLine[0] == 'Padded_CSR-Vector-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_pad_dp'] = curStats
                    elif curLine[0] == 'Padded_CSR-Vector-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_csr_vector_pad_dp_pcie'] = curStats
                    elif curLine[0] == 'ELLPACKR-SP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_ellpackr_sp'] = curStats
                    elif curLine[0] == 'ELLPACKR-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['spmv_ellpackr_dp'] = curStats

    return stats


def processStencil2D(folder):
    stats = {}
    filename = 'dev0_Stencil2D.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'SP_Sten2D':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['stencil'] = curStats
                    elif curLine[0] == 'DP_Sten2D':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['stencil_dp'] = curStats

    return stats


def processTriad(folder):
    stats = {}
    max = -1.0
    filename = 'dev0_Triad.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'TriadBdwth':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        if float(curStats['median']) > max:
                            max = float(curStats['median'])
                            stats = curStats
    else:
        print('Cannot find log file {}'.format(filename))

    return {'triad_bw': stats}


def processS3D(folder):
    stats = {}
    filename = 'dev0_S3D.log'
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath) as f:
            for line in f:
                curLine = line.split()
                if curLine:
                    if curLine[0] == 'S3D-SP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['s3d'] = curStats
                    elif curLine[0] == 'S3D-SP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['s3d_pcie'] = curStats
                    elif curLine[0] == 'S3D-DP':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['s3d_dp'] = curStats
                    elif curLine[0] == 'S3D-DP_PCIe':
                        curStats = {dataFormat[i]: curLine[i]
                                    for i in range(len(dataFormat))}
                        stats['s3d_dp_pcie'] = curStats

    return stats


def main(args):
    # Start stats collection
    stats = parseStats(args.inputFolder)

    if args.verbose:
        printStats(stats)

    output = args.output
    if output == '':
        output = os.path.dirname(args.inputFolder)

    if not args.disable:
        createSpreadsheet(args.inputFolder, output, stats)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Script to extract shoc execution-time data.',
                                    formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument(dest='inputFolder',
                        help='The path to a benchmark directory')

    parser.add_argument('--verbose', '-v',
                        action='store_true', default=False,
                        dest='verbose',
                        help='Print stats to terminal')

    parser.add_argument('--output', '-o',
                        default='',
                        dest='output',
                        help='Output results spreadsheet filepath containing only the median and variance')

    parser.add_argument('--all', '-a',
                        action='store_true', default=False,
                        dest='printAllRuns',
                        help='Write all runs statistics instead of only the median')

    parser.add_argument('--disable', '-d',
                        action='store_true', default=False,
                        dest='disable',
                        help='Disable spreadsheet output.')

    args = parser.parse_args()

    main(args)
