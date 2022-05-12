#!/usr/bin/python3
'''
**********************************************************************************
Description:  Extract schedule length stats from a plaidbench run.
Author:       Vang Thao
Created:      December 30, 2019
Last Update:  September 2020
**********************************************************************************

OUTPUT:
    This script takes in data from a plaidbench run and output a spreadsheet
    containing the average schedule length for each benchmark.
        Spreadsheet 1: schedule-length.xlsx

Requirements:
    - python3
    - pip3
    - openpyxl (sreadsheet module, installed using pip3)

HOW TO USE:
    1.) Run a plaidbench benchmarks with run-plaidbench.sh to generate a
        directory containing the results for the run.
    2.) Pass the path to the folder as an input to this script

Example:
    ./get-sched-length.py /home/tom/plaidbench-run-01
'''

import argparse
import logging
import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from readlogs import *

def parseStats(filePaths):
    # Get logger
    logger = logging.getLogger('parseStats')

    # Overall stats for the benchmark suite
    stats = {}

    for bench in filePaths:
        curStats = {}
        curStats['average'] = -1.0
        curStats['total'] = 0.0
        curStats['numRegions'] = 0
        curStats['maxLength'] = -1

        # First check if log file exists.
        if os.path.exists(filePaths[bench]):
            # Open log file if it exists.
            with open(filePaths[bench]) as file:
                # Read the whole log file
                # and split the scheduling
                # regions into a list
                log = file.read()
                blocks = split_blocks(log)

                # Iterate over each scheduling region
                for block in blocks:
                    events = keep_only_first_event(parse_events(block))

                    # Skip first pass because it isn't the
                    # final schedule
                    if ('PassFinished' in events.keys() and events['PassFinished']['num'] == 1):
                        continue

                    # First check if B&B is enabled because
                    # with B&B enabled, the final output will
                    # be different.
                    # If B&B is not enabled, check for
                    # schedule from heuristic.
                    schedLength = -1
                    if ('Enumerating' in events.keys()):
                        schedLength = events['BestResult']['length']
                    else:
                        schedLength = events['HeuristicResult']['length']

                    curStats['total'] += schedLength
                    curStats['numRegions'] += 1

                    if curStats['maxLength'] < schedLength:
                        curStats['maxLength'] = schedLength

            if curStats['numRegions'] != 0:
                curStats['average'] = curStats['total']/curStats['numRegions']

            stats[bench] = curStats

    return stats


def printStats(stats):
    overallRegions = 0
    overallTotal = 0
    overallAverage = -1
    overallMaxLength = -1

    for bench in stats:
        print('{} : Average: {:0.2f}, Max : {}, Regions: {}'.format(
            bench,
            stats[bench]['average'],
            stats[bench]['maxLength'],
            stats[bench]['numRegions']))

        overallRegions += stats[bench]['numRegions']
        overallTotal += stats[bench]['total']

        if overallMaxLength < stats[bench]['maxLength']:
            overallMaxLength = stats[bench]['maxLength']

    if overallRegions != 0:
        overallAverage = overallTotal / overallRegions

    print('Overall : Average: {:0.2f} Max : {}, Regions: {}'.format(
        overallAverage,
        overallMaxLength,
        overallRegions))


def createSpreadsheets(stats, outputFile):
    if 'xls' not in outputFile[-4:]:
        outputFile += '.xlsx'

    # Create new excel worksheet
    wb = Workbook()

    # Grab the active worksheet
    ws = wb.active

    # Insert title
    ws['A1'] = 'Benchmarks'
    ws['A1'].font = Font(bold=True)

    # Stats entry
    col = 'B'
    row = 1
    ws[col+str(row)] = 'Average Sched. Length'
    ws[chr(ord(col)+1)+str(row)] = 'Max Sched. Length'
    ws[chr(ord(col)+2)+str(row)] = 'Regions'

    overallRegions = 0
    overallTotal = 0
    overallAverage = -1
    overallMaxLength = -1
    row = 2
    for bench in stats:
        ws['A'+str(row)] = bench
        ws[col+str(row)] = stats[bench]['average']
        overallTotal += stats[bench]['total']

        ws[chr(ord(col)+1)+str(row)] = stats[bench]['maxLength']
        if overallMaxLength < stats[bench]['maxLength']:
            overallMaxLength = stats[bench]['maxLength']

        ws[chr(ord(col)+2)+str(row)] = stats[bench]['numRegions']
        overallRegions += stats[bench]['numRegions']
        row += 1

    if overallRegions != 0:
        overallAverage = overallTotal / overallRegions

    ws['A'+str(row)] = 'Overall'
    ws['A'+str(row)].font = Font(bold=True)
    ws[col+str(row)] = overallAverage
    ws[chr(ord(col)+1)+str(row)] = overallMaxLength
    ws[chr(ord(col)+2)+str(row)] = overallRegions

    wb.save(outputFile)


def main(args):
    if args.verbose:
        logging.basicConfig(level=logging.DEBUG)

    # Get filepaths for the selected benchmark suite
    filePaths = get_bench_log_paths(args.inputFolder, args.benchmark)

    # Start stats collection
    stats = parseStats(filePaths)

    # Print stats if enabled
    if args.verbose:
        printStats(stats)

    # Create spreadsheet
    if not args.disable:
        filename = ''
        if args.output is None:
            filename = os.path.dirname('schedule-length-' + args.inputFolder)
        else:
            filename = args.output

        createSpreadsheets(stats, filename)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Script to extract average schedule length.',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument(dest='inputFolder',
                        help='The path to a benchmark directory')

    parser.add_argument('--verbose', '-v',
                        action='store_true', default=False,
                        dest='verbose',
                        help='Print average schedule lengths to terminal')

    parser.add_argument('--output', '-o',
                        dest='output',
                        help='Output spreadsheet filepath')

    parser.add_argument('--disable', '-d',
                        action='store_true',
                        default=False, dest='disable',
                        help='Disable spreadsheet output.')

    parser.add_argument('--benchmark', '-b',
                        default='plaid',
                        choices=['plaid', 'shoc'],
                        dest='benchmark',
                        help='Select the benchmarking suite to parse for.')

    args = parser.parse_args()

    main(args)
