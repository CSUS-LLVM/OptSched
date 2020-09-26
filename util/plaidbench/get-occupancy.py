#!/usr/bin/python3
'''
**********************************************************************************
Description:  Extract occupancy stats from a plaidbench run.
Author:       Vang Thao
Created:      December 30, 2019
Last Update:  September 2020
**********************************************************************************

OUTPUT:
    This script takes in data from a plaidbench run and output a spreadsheet
    containing the average occupancy for each benchmark.
        Spreadsheet 1: occupancy.xlsx

Requirements:
    - python3
    - pip3
    - openpyxl (sreadsheet module, installed using pip3)
    - patch for LLVM to print out occupancy

HOW TO USE:
    1.) Run a plaidbench benchmarks with run-plaidbench.sh to generate a
        directory containing the results for the run.
    2.) Pass the path to the folder as an input to this script

Example:
    ./get-occupancy.py /home/tom/plaidbench-run-01
'''

import argparse
import logging
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from readlogs import *

def parseStats(filePaths):
    # Get logger
    logger = logging.getLogger('parseStats')

    # Overall stats for the benchmark suite
    stats = {}

    for bench in filePaths:
        curStats = {}
        curStats['average'] = -1.0
        curStats['total'] = 0.0
        curStats['numRegions'] = 0
        curStats['maxLength'] = -1
        curStats['numKernel'] = 0

        # First check if log file exists.
        if os.path.exists(filePaths[bench]):
            # Open log file if it exists.
            with open(filePaths[bench]) as file:
                for line in file:
                    # Match the line that contain occupancy stats
                    getOccupancyStats = OPT_RE_OCCUPANCY.match(line)
                    if (getOccupancyStats):
                        # Get the kernel name
                        kernelName = getOccupancyStats.group(1)

                        # Ignore these function
                        if (kernelName in OPT_IGNORE):
                            continue

                        # Get occupancy
                        occupancy = int(getOccupancyStats.group(2))

                        # Used for averaging
                        curStats['total'] += occupancy
                        curStats['numKernel'] += 1
                if curStats['numKernel'] != 0:
                    curStats['average'] = curStats['total'] / \
                        curStats['numKernel']
        else:
            print('Cannot find log file {}.'.format(filePaths[bench]))

        # Save stats
        stats[bench] = curStats
    return stats


def printStats(stats):
    total = 0.0
    kernel = 0
    for bench in stats:
        print('    {} : {:.2f}'.format(
            bench, stats[bench]['average']))
        total += stats[bench]['total']
        kernel += stats[bench]['numKernel']
    if kernel != 0:
        print('  Average: {:.2f}'.format(total/kernel))


def createSpreadsheets(stats, outputFile):
    if 'xls' not in outputFile[-4:]:
        outputFile += '.xlsx'

    # Create new excel worksheet
    wb = Workbook()

    # Grab the active worksheet
    ws = wb.active

    # Insert title and benchmark names
    ws['A1'] = 'Benchmarks'
    ws['A1'].font = Font(bold=True)

    # Stats entry
    col = 'B'
    row = 1
    ws[col+str(row)] = 'Occupancy'
    row = 2

    total = 0.0
    kernel = 0
    for bench in stats:
        ws['A' + str(row)] = bench
        ws[col+str(row)] = stats[bench]['average']
        total += stats[bench]['total']
        kernel += stats[bench]['numKernel']
        row += 1

    ws['A' + str(row)] = 'Average'
    ws['A' + str(row)].font = Font(bold=True)
    ws[col+str(row)] = total/kernel

    wb.save(outputFile)


def main(args):
    if args.verbose:
        logging.basicConfig(level=logging.DEBUG)

    # Get filepaths for the selected benchmark suite
    filePaths = get_bench_log_paths(args.inputFolder, args.benchmark)

    # Start stats collection
    stats = parseStats(filePaths)

    # Print stats if enabled
    if args.verbose:
        printStats(stats)

    # Create spreadsheet
    if not args.disable:
        filename = ''
        if args.output is None:
            filename = os.path.dirname('occupancy-' + args.inputFolder)
        else:
            filename = args.output

        createSpreadsheets(stats, filename)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Script to extract occupancy data. \
                                     Requires patch to print occupancy.',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument(dest='inputFolder',
                        help='The path to a benchmark directory')

    parser.add_argument('--verbose', '-v',
                        action='store_true', default=False,
                        dest='verbose',
                        help='Print average occupancy to terminal')

    parser.add_argument('--output', '-o', dest='output',
                        help='Output spreadsheet filepath')

    parser.add_argument('--disable', '-d',
                        action='store_true', default=False,
                        dest='disable',
                        help='Disable spreadsheet output.')

    parser.add_argument('--benchmark', '-b',
                        default='plaid',
                        choices=['plaid', 'shoc'],
                        dest='benchmark',
                        help='Select the benchmarking suite to parse for.')

    args = parser.parse_args()

    main(args)
